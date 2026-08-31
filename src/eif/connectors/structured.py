"""Connectors for structured/tabular formats: JSON, CSV, Excel.

These normalize structured data into text Evidence with a compact, model-friendly
rendering while preserving the raw values in metadata where practical. Numeric
tables are additionally summarized (row count, column headers) so downstream
extractors can pick up measurements deterministically.
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from typing import Any

from ..domain import Evidence
from ..domain.enums import Modality, SourceType
from ..exceptions import ConnectorError
from .base import EIFConnector
from .text import _is_existing_file


class JsonConnector(EIFConnector):
    """Loads ``.json`` files and in-memory dict/list payloads."""

    modality = Modality.JSON
    source_type = SourceType.API

    def can_handle(self, source: Any) -> bool:
        if _is_existing_file(source, {".json"}):
            return True
        return isinstance(source, dict | list)

    def load(self, source: Any) -> list[Evidence]:
        if _is_existing_file(source, {".json"}):
            path = Path(source)
            raw = self.read_file_bytes(path).decode("utf-8", errors="replace")
            try:
                data = json.loads(raw)
            except json.JSONDecodeError as exc:
                raise ConnectorError(f"Invalid JSON in {path}: {exc}") from exc
            name = path.name
        else:
            data = source
            name = "inline-json"
        pretty = json.dumps(data, indent=2, sort_keys=True, default=str)
        return [
            self.make_text_evidence(
                pretty, source=name, modality=Modality.JSON, mime_type="application/json"
            )
        ]


class CsvConnector(EIFConnector):
    """Loads ``.csv`` / ``.tsv`` files into a text rendering + summary."""

    modality = Modality.TABLE
    source_type = SourceType.FILE

    def can_handle(self, source: Any) -> bool:
        return _is_existing_file(source, {".csv", ".tsv"})

    def load(self, source: Any) -> list[Evidence]:
        path = Path(source)
        raw = self.read_file_bytes(path).decode("utf-8", errors="replace")
        delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
        reader = csv.reader(io.StringIO(raw), delimiter=delimiter)
        rows = list(reader)
        if not rows:
            raise ConnectorError(f"CSV file is empty: {path}")
        headers = rows[0]
        body_rows = rows[1:]
        rendered = _render_table(headers, body_rows)
        metadata = {
            "rows": str(len(body_rows)),
            "columns": ",".join(headers),
        }
        return [
            self.make_text_evidence(
                rendered,
                source=path.name,
                modality=Modality.TABLE,
                mime_type="text/csv",
                metadata=metadata,
            )
        ]


class ExcelConnector(EIFConnector):
    """Loads ``.xlsx`` files (requires the optional ``openpyxl`` dependency)."""

    modality = Modality.TABLE
    source_type = SourceType.FILE

    def can_handle(self, source: Any) -> bool:
        return _is_existing_file(source, {".xlsx", ".xlsm"})

    def load(self, source: Any) -> list[Evidence]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - optional dep
            raise ConnectorError(
                "Reading Excel files requires the optional 'openpyxl' dependency. "
                "Install with: pip install 'economic-intelligence-framework[excel]'"
            ) from exc
        path = Path(source)
        # Validate size/mime via the shared reader, then let openpyxl parse the path.
        self.read_file_bytes(path)
        wb = load_workbook(path, read_only=True, data_only=True)
        evidences: list[Evidence] = []
        for sheet in wb.worksheets:
            rows = [
                [("" if cell is None else str(cell)) for cell in row]
                for row in sheet.iter_rows(values_only=True)
            ]
            if not rows:
                continue
            headers = rows[0]
            rendered = _render_table(headers, rows[1:])
            evidences.append(
                self.make_text_evidence(
                    rendered,
                    source=f"{path.name}::{sheet.title}",
                    modality=Modality.TABLE,
                    mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    metadata={"sheet": sheet.title, "rows": str(len(rows) - 1)},
                )
            )
        wb.close()
        if not evidences:
            raise ConnectorError(f"No readable sheets in {path}")
        return evidences


def _render_table(headers: list[str], rows: list[list[str]], *, max_rows: int = 200) -> str:
    """Render a table as pipe-delimited text with a truncation note."""
    lines = [" | ".join(headers), " | ".join(["---"] * len(headers))]
    for row in rows[:max_rows]:
        lines.append(" | ".join(row))
    if len(rows) > max_rows:
        lines.append(f"... ({len(rows) - max_rows} more rows omitted)")
    return "\n".join(lines)
